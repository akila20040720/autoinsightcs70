from __future__ import annotations

import argparse
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def normalize_text(value: str | None) -> str:
    text = (value or "").strip().lower()
    text = text.replace("&", " and ").replace("-", " ")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_year_bounds(value: object) -> tuple[int | None, int | None]:
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        year = int(value)
        return year, year

    text = str(value).strip()
    if not text:
        return None, None

    years = [int(match) for match in re.findall(r"(?:19|20)\d{2}", text)]
    if years:
        if len(years) == 1:
            return years[0], years[0]
        return min(years), max(years)

    match = re.fullmatch(r"(\d{4})\s*[-/]\s*(\d{4})", text)
    if match:
        start, end = int(match.group(1)), int(match.group(2))
        return min(start, end), max(start, end)

    return None, None


def parse_currency_value(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(",", "")
    match = re.search(r"(\d+(?:\.\d+)?)", cleaned)
    if not match:
        return None
    try:
        return int(float(match.group(1)))
    except (TypeError, ValueError):
        return None


def parse_avg_price_mileage(value: object) -> tuple[int | None, int | None]:
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    parts = [part.strip() for part in text.split("|")]
    price = parse_currency_value(parts[0]) if parts else None
    mileage = parse_currency_value(parts[1]) if len(parts) > 1 else None
    return price, mileage


def build_trend_labels(header_rows: list[tuple[object, ...]]) -> list[str]:
    years_row = list(header_rows[1]) if len(header_rows) > 1 else []
    months_row = list(header_rows[2]) if len(header_rows) > 2 else []
    labels: list[str] = []
    current_year: str | None = None

    for index in range(3, 9):
        if index < len(years_row) and years_row[index]:
            current_year = str(years_row[index]).strip()
        month_label = (
            str(months_row[index]).strip()
            if index < len(months_row) and months_row[index]
            else f"Point {index - 2}"
        )
        month_label = month_label.replace("MARCH", "MAR").replace("APRIL", "APR")
        month_label = re.sub(r"\s+", " ", month_label)
        labels.append(f"{month_label} {current_year}".strip())

    return labels


def extract_trend_points(row: tuple[object, ...], labels: list[str]) -> list[dict[str, object]]:
    points: list[dict[str, object]] = []
    for offset, index in enumerate(range(3, 9)):
        if index >= len(row):
            continue
        value = parse_currency_value(row[index])
        if value is None:
            continue
        label = labels[offset] if offset < len(labels) else f"Point {offset + 1}"
        points.append(
            {
                "label": label,
                "valueLkr": value,
                "predicted": ("predicted" in label.lower()) or ("apr" in label.lower()),
            }
        )
    return points


def workbook_to_mongo_docs(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    docs: list[dict[str, Any]] = []
    exported_at = datetime.now(tz=UTC).isoformat()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_rows = list(worksheet.iter_rows(min_row=1, max_row=3, values_only=True))
        trend_labels = build_trend_labels(header_rows)

        for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
            if not row:
                continue

            make = row[0] if len(row) > 0 else None
            model = row[1] if len(row) > 1 else None
            year_value = row[2] if len(row) > 2 else None

            if not make or not model:
                continue

            make_text = str(make).strip()
            model_text = str(model).strip()
            if not make_text or not model_text:
                continue
            if make_text.lower() == "make" or model_text.lower() == "model":
                continue

            year_min, year_max = parse_year_bounds(year_value)
            trend_points = extract_trend_points(row, trend_labels)
            avg_price_lkr, avg_mileage = parse_avg_price_mileage(row[10] if len(row) > 10 else None)
            previous_month = parse_currency_value(row[7] if len(row) > 7 else None)
            next_week = parse_currency_value(row[9] if len(row) > 9 else None)

            docs.append(
                {
                    "rowId": f"{sheet_name}:{row_index}",
                    "sourceSheet": sheet_name,
                    "sourceRow": row_index,
                    "vehicleType": "Car",
                    "make": make_text,
                    "model": model_text,
                    "makeNorm": normalize_text(make_text),
                    "modelNorm": normalize_text(model_text),
                    "yearMin": year_min,
                    "yearMax": year_max,
                    "market_analysis": {
                        "previousMonthPriceLkr": previous_month,
                        "nextWeekPriceLkr": next_week,
                        "avgPriceLkr": avg_price_lkr,
                        "avgMileage": avg_mileage,
                        "priceTrend": trend_points,
                    },
                    "updatedAt": exported_at,
                }
            )

    return docs


def main() -> int:
    parser = argparse.ArgumentParser(description="Export model template workbook into Mongo-ready JSON")
    parser.add_argument("--xlsx", required=True, help="Path to template xlsx file")
    parser.add_argument("--out", required=True, help="Path to output json file")
    args = parser.parse_args()

    workbook_path = Path(args.xlsx).resolve()
    output_path = Path(args.out).resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    docs = workbook_to_mongo_docs(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": {
            "sourceWorkbook": str(workbook_path),
            "collection": "model_reference_rows",
            "documents": len(docs),
            "exportedAt": datetime.now(tz=UTC).isoformat(),
        },
        "documents": docs,
    }

    output_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")

    mongo_array_path = output_path.with_name(f"{output_path.stem}.mongoimport.json")
    mongo_array_path.write_text(json.dumps(docs, ensure_ascii=True, indent=2), encoding="utf-8")

    print(f"Exported {len(docs)} documents -> {output_path}")
    print(f"Mongo import array -> {mongo_array_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
