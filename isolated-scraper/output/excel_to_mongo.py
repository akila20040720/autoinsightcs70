#!/usr/bin/env python3
"""
Convert an Excel file to JSON and optionally insert records into MongoDB.

Usage:
1) Fill MONGO_URI with your connection string.
2) Run: python excel_to_mongo.py
"""

from __future__ import annotations

import json
import math
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, List

from openpyxl import load_workbook

# ==== EDIT THESE VALUES ====
EXCEL_FILE = Path("isolated-scraper/output/Template_for_Model_Filled (1).xlsx")
JSON_FILE = Path("isolated-scraper/output/Template_for_Model_Filled.json")
MONGO_URI = ""
MONGO_DB = "autoinsight"
MONGO_COLLECTION = "vehicle_listings"
SHEET_NAME = None  # Example: "Sheet1". Keep None to use the first sheet.
# ===========================


def clean_key(value: Any, index: int) -> str:
    if value is None:
        return f"field_{index + 1}"
    key = str(value).strip()
    return key if key else f"field_{index + 1}"


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def keyify(value: str) -> str:
    text = value.strip().lower()
    text = text.replace("(", " ").replace(")", " ")
    text = text.replace("-", " ").replace("_", " ")
    return " ".join(text.split())


def parse_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return default
    cleaned = re.sub(r"[^0-9]", "", text)
    return int(cleaned) if cleaned else default


def parse_price_lkr(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        number = float(value)
        # If the sheet stores values in millions, convert to full LKR.
        if 0 < number < 10_000:
            return int(number * 1_000_000)
        return int(number)
    text = str(value).strip().lower().replace(",", "")
    if not text:
        return 0
    if "million" in text or " mn" in text or text.endswith("m"):
        digits = re.sub(r"[^0-9.]", "", text)
        return int(float(digits) * 1_000_000) if digits else 0
    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else 0


def normalize_condition(value: Any) -> str:
    text = str(value or "").strip().lower()
    if "brand new" in text or "unregistered" in text or text == "new":
        return "Brand New"
    if "recondition" in text:
        return "Recondition"
    return "Used"


def normalize_vehicle_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"car", "cars"}:
        return "Car"
    if text in {"van", "vans"}:
        return "Van"
    if text in {"suv", "suvs"}:
        return "SUV"
    if text in {"motorbike", "motorbikes", "bike", "bikes"}:
        return "Motorbike"
    return "Car"


def normalize_listed_at(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    text = str(value or "").strip()
    if not text:
        return datetime.now().isoformat()
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.isoformat()
    except ValueError:
        return datetime.now().isoformat()


def get_field(raw: dict[str, Any], aliases: list[str], default: Any = "") -> Any:
    for alias in aliases:
        if alias in raw and raw[alias] not in (None, ""):
            return raw[alias]
    return default


def to_frontend_vehicle(raw: dict[str, Any], index: int) -> dict[str, Any]:
    vehicle_url = str(get_field(raw, ["vehicle url", "vehicleurl", "url", "link"], "")).strip()
    make = str(get_field(raw, ["make", "brand"], "")).strip()
    model = str(get_field(raw, ["model"], "")).strip()
    year = parse_int(get_field(raw, ["year"], 0), 0)
    price_lkr = parse_price_lkr(get_field(raw, ["pricelkr", "price", "rawprice"], 0))
    mileage = parse_int(get_field(raw, ["mileage", "milleage", "km"], 0), 0)
    district = str(get_field(raw, ["district", "city", "location"], "")).strip()
    published_date = str(get_field(raw, ["published date", "publisheddate", "date"], "")).strip()
    listed_at = normalize_listed_at(get_field(raw, ["listedat", "published date", "publisheddate", "date"], ""))
    condition = normalize_condition(get_field(raw, ["condition", "title", "status"], "Used"))
    vehicle_type = normalize_vehicle_type(get_field(raw, ["vehicle type", "vehicletype", "type"], "Car"))
    image_url = str(get_field(raw, ["imageurl", "image", "image url"], "")).strip()

    generated_id = f"listing-{index + 1}"
    if make or model or year:
        generated_id = f"{make}-{model}-{year}-{index + 1}".strip("-").replace(" ", "-").lower()

    vehicle_id = str(get_field(raw, ["id", "listing id", "listingid"], "")).strip() or vehicle_url or generated_id

    return {
        "id": vehicle_id,
        "vehicleType": vehicle_type,
        "make": make,
        "model": model,
        "year": year,
        "priceLkr": price_lkr,
        "priceMillion": round(price_lkr / 1_000_000, 2),
        "mileage": mileage,
        "district": district,
        "publishedDate": published_date,
        "listedAt": listed_at,
        "vehicleUrl": vehicle_url,
        "condition": condition,
        "imageUrl": image_url or None,
        "validationStatus": "unmatched",
        "confidence": 0,
        "matchedModelRowId": None,
        "isActive": True,
    }


def row_to_raw_doc(headers: list[str], row: Any) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for i, key in enumerate(headers):
        cell_value = row[i] if i < len(row) else None
        raw[keyify(key)] = normalize_value(cell_value)
    return raw


def excel_to_documents(excel_path: Path, sheet_name: str | None = None) -> List[dict[str, Any]]:
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

        rows = sheet.iter_rows(values_only=True)
        headers_row = next(rows, None)
        if headers_row is None:
            return []

        headers = [clean_key(v, i) for i, v in enumerate(headers_row)]

        documents: List[dict[str, Any]] = []
        for index, row in enumerate(rows):
            if row is None:
                continue

            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            raw_doc = row_to_raw_doc(headers, row)
            doc = to_frontend_vehicle(raw_doc, index)
            documents.append(doc)

        return documents
    finally:
        workbook.close()


def save_json(documents: List[dict[str, Any]], json_path: Path) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(documents, f, ensure_ascii=False, indent=2)


def insert_into_mongo(uri: str, db_name: str, collection_name: str, documents: List[dict[str, Any]]) -> int:
    from pymongo import MongoClient

    client = MongoClient(uri, serverSelectionTimeoutMS=10000)
    try:
        client.admin.command("ping")
        collection = client[db_name][collection_name]
        if not documents:
            return 0
        result = collection.insert_many(documents)
        return len(result.inserted_ids)
    finally:
        client.close()


def main() -> None:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    documents = excel_to_documents(EXCEL_FILE, SHEET_NAME)
    save_json(documents, JSON_FILE)

    print(f"Converted to JSON: {JSON_FILE}")
    print(f"Total records: {len(documents)}")

    if not MONGO_URI.strip():
        print("Mongo insert skipped: MONGO_URI is empty.")
        print("Set MONGO_URI in this file, then run again to insert into MongoDB.")
        return

    inserted_count = insert_into_mongo(MONGO_URI, MONGO_DB, MONGO_COLLECTION, documents)
    print(
        f"Inserted {inserted_count} records into MongoDB: "
        f"{MONGO_DB}.{MONGO_COLLECTION}"
    )


if __name__ == "__main__":
    main()
