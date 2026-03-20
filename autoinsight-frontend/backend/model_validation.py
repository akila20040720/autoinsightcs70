from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def normalize_text(value: str | None) -> str:
    text = (value or "").strip().lower()
    text = text.replace("&", " and ").replace("-", " ")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _token_set(value: str | None) -> set[str]:
    return {token for token in normalize_text(value).split() if token}


def _parse_year_bounds(value: object) -> tuple[int | None, int | None]:
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        year = int(value)
        return year, year

    text = str(value).strip()
    if not text:
        return None, None

    years = [int(match) for match in re.findall(r"(?:19|20)\d{2}", text)]
    if years:
        if len(years) == 1:
            return years[0], years[0]
        return min(years), max(years)

    match = re.fullmatch(r"(\d{4})\s*[-/]\s*(\d{4})", text)
    if match:
        start, end = int(match.group(1)), int(match.group(2))
        return min(start, end), max(start, end)

    return None, None


@dataclass(frozen=True)
class ModelReferenceRow:
    row_id: str
    vehicle_type: str | None
    make: str
    model: str
    year_min: int | None
    year_max: int | None
    previous_month_price_lkr: int | None
    next_week_price_lkr: int | None
    avg_price_lkr: int | None
    avg_mileage: int | None
    trend_points: list[dict[str, object]]

    @property
    def normalized_make(self) -> str:
        return normalize_text(self.make)

    @property
    def normalized_model(self) -> str:
        return normalize_text(self.model)


class ModelValidator:
    def __init__(self, rows: Iterable[ModelReferenceRow]):
        self.rows = list(rows)
        self.rows_by_make: dict[str, list[ModelReferenceRow]] = {}
        for row in self.rows:
            self.rows_by_make.setdefault(row.normalized_make, []).append(row)

    @classmethod
    def from_workbook(cls, workbook_path: Path) -> "ModelValidator":
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        rows: list[ModelReferenceRow] = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_rows = list(worksheet.iter_rows(min_row=1, max_row=3, values_only=True))
            trend_labels = _build_trend_labels(header_rows)
            for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
                if not row:
                    continue
                make = row[0] if len(row) > 0 else None
                model = row[1] if len(row) > 1 else None
                year_value = row[2] if len(row) > 2 else None

                if not make or not model:
                    continue

                make_text = str(make).strip()
                model_text = str(model).strip()
                if not make_text or not model_text:
                    continue

                if make_text.lower() == "make" or model_text.lower() == "model":
                    continue

                year_min, year_max = _parse_year_bounds(year_value)
                trend_points = _extract_trend_points(row, trend_labels)
                avg_price_lkr, avg_mileage = _parse_avg_price_mileage(row[10] if len(row) > 10 else None)
                rows.append(
                    ModelReferenceRow(
                        row_id=f"{sheet_name}:{row_index}",
                        vehicle_type="Car",
                        make=make_text,
                        model=model_text,
                        year_min=year_min,
                        year_max=year_max,
                        previous_month_price_lkr=_parse_currency_value(row[7] if len(row) > 7 else None),
                        next_week_price_lkr=_parse_currency_value(row[9] if len(row) > 9 else None),
                        avg_price_lkr=avg_price_lkr,
                        avg_mileage=avg_mileage,
                        trend_points=trend_points,
                    )
                )

        return cls(rows)

    def validate_listing(self, listing: dict[str, object]) -> dict[str, object]:
        make = normalize_text(str(listing.get("make") or ""))
        model = normalize_text(str(listing.get("model") or ""))
        vehicle_type = normalize_text(str(listing.get("vehicleType") or ""))
        year = listing.get("year")

        try:
            year_value = int(year) if year is not None else None
        except (TypeError, ValueError):
            year_value = None

        if not make or not model:
            return {
                "validation_status": "unmatched",
                "confidence": 0.0,
                "matched_model_row_id": None,
            }

        candidates = self.rows_by_make.get(make, [])
        best_row: ModelReferenceRow | None = None
        best_score = 0.0

        listing_tokens = _token_set(model)

        for row in candidates:
            row_model = row.normalized_model
            row_tokens = _token_set(row_model)
            if not row_model:
                continue

            score = 0.0
            if row_model == model:
                score += 0.72
            elif row_model in model or model in row_model:
                score += 0.62
            else:
                overlap = len(listing_tokens & row_tokens)
                if overlap == 0:
                    continue
                coverage = overlap / max(len(row_tokens), 1)
                if coverage < 0.5:
                    continue
                score += 0.35 + (coverage * 0.2)

            if row.vehicle_type:
                row_vehicle_type = normalize_text(row.vehicle_type)
                if vehicle_type and row_vehicle_type == vehicle_type:
                    score += 0.08
                elif vehicle_type and row_vehicle_type != vehicle_type:
                    score -= 0.12

            if year_value is not None and row.year_min is not None:
                if row.year_min <= year_value <= (row.year_max or row.year_min):
                    score += 0.2
                elif abs(year_value - row.year_min) <= 1:
                    score += 0.08
                else:
                    score -= 0.18
            else:
                score += 0.03

            if score > best_score:
                best_score = score
                best_row = row

        confidence = round(max(0.0, min(best_score, 0.99)), 2)
        if best_row is None or confidence < 0.45:
            return {
                "validation_status": "unmatched",
                "confidence": confidence,
                "matched_model_row_id": None,
            }

        status = "validated" if confidence >= 0.8 else "partial"
        return {
            "validation_status": status,
            "confidence": confidence,
            "matched_model_row_id": best_row.row_id,
            "market_analysis": {
                "previousMonthPriceLkr": best_row.previous_month_price_lkr,
                "nextWeekPriceLkr": best_row.next_week_price_lkr,
                "avgPriceLkr": best_row.avg_price_lkr,
                "avgMileage": best_row.avg_mileage,
                "priceTrend": best_row.trend_points,
            },
        }


def _parse_currency_value(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(",", "")
    match = re.search(r"(\d+(?:\.\d+)?)", cleaned)
    if not match:
        return None
    try:
        return int(float(match.group(1)))
    except (TypeError, ValueError):
        return None


def _parse_avg_price_mileage(value: object) -> tuple[int | None, int | None]:
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    parts = [part.strip() for part in text.split("|")]
    price = _parse_currency_value(parts[0]) if parts else None
    mileage = _parse_currency_value(parts[1]) if len(parts) > 1 else None
    return price, mileage


def _build_trend_labels(header_rows: list[tuple[object, ...]]) -> list[str]:
    years_row = list(header_rows[1]) if len(header_rows) > 1 else []
    months_row = list(header_rows[2]) if len(header_rows) > 2 else []
    labels: list[str] = []
    current_year: str | None = None

    for index in range(3, 9):
        if index < len(years_row) and years_row[index]:
            current_year = str(years_row[index]).strip()
        month_label = str(months_row[index]).strip() if index < len(months_row) and months_row[index] else f"Point {index - 2}"
        month_label = month_label.replace("MARCH", "MAR").replace("APRIL", "APR")
        month_label = re.sub(r"\s+", " ", month_label)
        labels.append(f"{month_label} {current_year}".strip())

    return labels


def _extract_trend_points(row: tuple[object, ...], labels: list[str]) -> list[dict[str, object]]:
    points: list[dict[str, object]] = []
    for offset, index in enumerate(range(3, 9)):
        if index >= len(row):
            continue
        value = _parse_currency_value(row[index])
        if value is None:
            continue
        label = labels[offset] if offset < len(labels) else f"Point {offset + 1}"
        points.append(
            {
                "label": label,
                "valueLkr": value,
                "predicted": "predicted" in label.lower() or "apr" in label.lower(),
            }
        )
    return points
